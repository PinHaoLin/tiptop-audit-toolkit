from datetime import datetime, timedelta
import shutil
import os
import re
import openpyxl
import sys

def get_project_root():
    """ 取得專案根目錄，相容 PyInstaller 打包模式 """
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        current_script_dir = os.path.dirname(os.path.abspath(__file__))
        if os.path.basename(current_script_dir) == "backend":
            return os.path.abspath(os.path.join(current_script_dir, ".."))
        return current_script_dir

def is_red_color(cell):
    """ 檢查儲存格字型或背景填滿是否為紅色 """
    if not cell:
        return False
    red_rgbs = {"FFFF0000", "FF0000"}
    if cell.font and cell.font.color and cell.font.color.rgb in red_rgbs:
        return True
    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb in red_rgbs:
        return True
    return False

def is_file_matching_excel(pure_base_name, excel_f_val):
    """
    比對實體檔案基底名與 Excel 登記的程式代號。
    支援移除常見語系與客製後綴（如 _std, _tw, _cn 等）後的精確比對。
    """
    pb_name = pure_base_name.strip().lower()
    f_val = excel_f_val.strip().lower()
    
    if not pb_name or not f_val:
        return False
    if pb_name == f_val:
        return True
        
    suffixes = ["_std", "_tw", "_cn", "_en", "_zh", "_std_tw", "_0"]
    for suffix in suffixes:
        if pb_name.endswith(suffix) and pb_name[:-len(suffix)] == f_val:
            return True
        if f_val.endswith(suffix) and f_val[:-len(suffix)] == pb_name:
            return True
        if f_val + suffix == pb_name:
            return True
            
    return False

def backup_and_rename_by_latest_folder(base_dir_ignored=None, excel_source_path=None):
    """ 【第一階段】自動複製上週 .docx 查核表與複製DataCenter中 程式修改紀錄Excel 檔案 """
    project_root_dir = get_project_root()

    print("\n============================================================", flush=True)
    print(f"核心工作區（專案根目錄）已鎖定：{project_root_dir}", flush=True)

    aud_dir_path = os.path.join(project_root_dir, "aud")
    aud_warning = False

    if not os.path.exists(aud_dir_path):
        print("【提醒】未在專案目錄下偵測到 'aud' 資料夾！", flush=True)
        print(f"    請確認是否已將 Linux 查核腳本(z_chk_all.sh) 放置於：{aud_dir_path}", flush=True)
        aud_warning = True
    else:
        print(f"偵測到本地核心 'aud' 資料夾已存在 (路徑: {aud_dir_path})，符合執行環境規範，繼續往下執行...", flush=True)

    HARDCODED_BASE_DIR = os.path.join(project_root_dir, "程式修改週覆核記錄(暫存)")

    today = datetime.now()
    today_str = today.strftime("%Y%m%d")

    # 精準計算 7 天前與 1 天前
    date_7d_ago = (today - timedelta(days=7)).strftime("%Y%m%d")
    date_1d_ago = (today - timedelta(days=1)).strftime("%Y%m%d")

    if not os.path.exists(HARDCODED_BASE_DIR):
        print(f"偵測到未建立暫存母資料夾，正在為您動態新建 -> {HARDCODED_BASE_DIR}", flush=True)
        os.makedirs(HARDCODED_BASE_DIR, exist_ok=True)
    else:
        print("暫存母資料夾已存在，自動略過新建步驟，繼續處理今日覆核內容...", flush=True)

    new_folder_name = f"程式修改週覆核記錄-{today_str}"
    new_folder_path = os.path.join(HARDCODED_BASE_DIR, new_folder_name)

    # 尋找上週舊資料夾（嚴格排除今日資料夾）
    last_docx_path = None
    if os.path.exists(HARDCODED_BASE_DIR):
        existing_folders = [
            d for d in os.listdir(HARDCODED_BASE_DIR)
            if os.path.isdir(os.path.join(HARDCODED_BASE_DIR, d))
            and d.startswith("程式修改週覆核記錄-")
            and d != new_folder_name
        ]
        existing_folders.sort()

        for folder in reversed(existing_folders):
            folder_full_path = os.path.join(HARDCODED_BASE_DIR, folder)
            docx_files = [f for f in os.listdir(folder_full_path) if f.endswith(".docx") and "資訊處理定期覆核表" in f]
            if docx_files:
                last_docx_path = os.path.join(folder_full_path, docx_files[0])
                print(f"成功追蹤到真正的歷史查核表範本：{last_docx_path}", flush=True)
                break

    print(f"今日目標獨立資料夾: {new_folder_path}", flush=True)

    if os.path.exists(new_folder_path):
        print("偵測到今日資料夾已存在，執行完整刷新（刪除舊殘留）...", flush=True)
        try:
            shutil.rmtree(new_folder_path)
        except Exception as e:
            print(f"刪除舊資料夾失敗 (可能檔案被開啟): {e}", flush=True)

    os.makedirs(new_folder_path, exist_ok=True)

    attachment_two_name = f"附件二 {date_7d_ago}~{date_1d_ago}異動的4gl_4fd檔案.txt"
    attachment_three_name = f"附件三 {date_7d_ago}~{date_1d_ago}異動的rpt_xml檔案.txt"
    docx_name = f"MMF-MIS-068-0 資訊處理定期覆核表-{today_str}.docx"
    target_docx_full_path = os.path.join(new_folder_path, docx_name)
    attachment_three_content = (
        f"D:\\aud>type {today_str}_tiptop_rpt.csv\n\n"
        f"D:\\aud>type {today_str}_topcust_rpt.csv\n\n"
        f"D:\\aud>pause\n"
        f"請按任意鍵繼續 . . .\n\n\n"
        f"D:\\aud>type {today_str}_tiptop_xml.csv\n\n"
        f"D:\\aud>type {today_str}_topcust_xml.csv\n\n"
        f"D:\\aud>pause\n"
        f"請按任意鍵繼續 . . .\n\n\n"
    )

    try:
        if last_docx_path and os.path.exists(last_docx_path):
            shutil.copy2(last_docx_path, target_docx_full_path)
            print(f"已成功自歷史目錄承接並複製查核表 -> {docx_name}", flush=True)
        else:
            print("警告：未找到任何歷史查核表 (.docx)，系統將自動建立基礎檔案防呆。", flush=True)
            with open(target_docx_full_path, 'w', encoding='utf-8') as f: pass

        with open(os.path.join(new_folder_path, attachment_two_name), 'w', encoding='utf-8') as f: f.write("")
        with open(os.path.join(new_folder_path, attachment_three_name), 'w', encoding='utf-8') as f: f.write(attachment_three_content)

        if excel_source_path:
            print(f"正在從網碟來源複製原始 Excel...\n路徑: {excel_source_path}", flush=True)
            copy_and_rename_file(excel_source_path, new_folder_path, "02.GP程式修改記錄vNEW.xlsx")
        else:
            print("警告：未收到來自前端或後端的 EXCEL_SOURCE 路徑參數，跳過 Excel 複製。", flush=True)

        modify_sh_script(filename="z_chk_all.sh")

        print("今日本機獨立工作區初始化成功！", flush=True)
        return new_folder_path, attachment_two_name, attachment_three_name, today_str, aud_warning
    except Exception as e:
        print(f"初始化檔案或複製查核表失敗: {e}", flush=True)
        return new_folder_path, attachment_two_name, attachment_three_name, today_str, aud_warning

def copy_and_rename_file(source_path, target_directory, base_filename):
    """ 從網碟將最新的 Excel 複製備份至本機今日資料夾 """
    if not target_directory or not os.path.exists(source_path):
        print(f"Excel 來源路徑不存在或無權限: {source_path}", flush=True)
        return None

    today_str = datetime.now().strftime("%Y%m%d")
    new_name = base_filename.replace("NEW", today_str)
    destination_path = os.path.join(target_directory, new_name)
    try:
        shutil.copy2(source_path, destination_path)
        print(f"Excel 複製成功 -> {destination_path}", flush=True)
        return destination_path
    except Exception as e:
        print(f"複製 Excel 失敗: {e}", flush=True)
        return None

def modify_sh_script(filename="z_chk_all.sh", *args, **kwargs):
    """ 統一對齊至核心 aud 目錄：修改與 z_chk_all.sh 同資料夾下的腳本日期 """
    project_root_dir = get_project_root()
    aud_dir_path = os.path.join(project_root_dir, "aud")
    full_path = os.path.join(aud_dir_path, filename)

    if not os.path.exists(full_path):
        print(f"找不到指令腳本: {full_path}", flush=True)
        return

    today = datetime.now()
    yy_mm_dd_7 = (today - timedelta(days=7)).strftime("%y%m%d")
    yy_mm_dd_1 = (today - timedelta(days=1)).strftime("%y%m%d")
    try:
        with open(full_path, 'r', encoding='utf-8') as f: lines = f.readlines()
        updated_lines = []
        touch_count = 0
        for line in lines:
            if line.strip().startswith("touch -t"):
                touch_count += 1
                if touch_count == 1: updated_lines.append(line)
                elif touch_count == 2: updated_lines.append(re.sub(r"touch -t \d{6}(\d{4} end)", f"touch -t {yy_mm_dd_1}\\1", line))
                elif touch_count == 3: updated_lines.append(re.sub(r"touch -t \d{6}(\d{4} start)", f"touch -t {yy_mm_dd_7}\\1", line))
                elif touch_count == 4: updated_lines.append(re.sub(r"touch -t \d{6}(\d{4} end)", f"touch -t {yy_mm_dd_1}\\1", line))
            else: updated_lines.append(line)
        with open(full_path, 'w', encoding='utf-8', newline='\n') as f: f.writelines(updated_lines)
        print(f"已成功自動改寫 Linux 腳本日期區間 -> {full_path}", flush=True)
    except Exception as e:
        print(f"修改 SH 腳本失敗: {e}", flush=True)

def modify_bat_scripts(local_dir=None, *args, **kwargs):
    """ 統一對齊至核心 aud 目錄：直接在 aud 底下重構或輸出 Bat 腳本 """
    project_root_dir = local_dir or get_project_root()
    aud_dir_path = os.path.join(project_root_dir, "aud")

    today = datetime.now()
    target_today_str = today.strftime("%Y%m%d")
    target_7d_slashes = (today - timedelta(days=7)).strftime("%Y/%m/%d")

    # 依據 bat.txt 範本完整重構內容
    rpt_content = (
        f"@echo on\n\n"
        f"forfiles /P D:\\tiptop_cr\\topprod\\tiptop\\ /M *.rpt /S /D +{target_7d_slashes} /c \"cmd /c echo @fdate @ftime @path\" > {target_today_str}_tiptop_rpt.csv\n\n"
        f"forfiles /P D:\\tiptop_cr\\topprod\\topcust\\ /M *.rpt /S /D +{target_7d_slashes} /c \"cmd /c echo @fdate @ftime @path\" > {target_today_str}_topcust_rpt.csv\n\n"
        f"pause\n\n"
        f"type {target_today_str}_tiptop_rpt.csv\n\n"
        f"type {target_today_str}_topcust_rpt.csv\n\n"
        f"pause"
    )

    xml_content = (
        f"@echo on\n\n"
        f"forfiles /P D:\\tiptop_cr\\topprod\\tiptop\\ /M *.xml /S /D +{target_7d_slashes} /c \"cmd /c echo @fdate @ftime @path\" > {target_today_str}_tiptop_xml.csv\n\n"
        f"forfiles /P D:\\tiptop_cr\\topprod\\topcust\\ /M *.xml /S /D +{target_7d_slashes} /c \"cmd /c echo @fdate @ftime @path\" > {target_today_str}_topcust_xml.csv\n\n"
        f"pause\n\n"
        f"type {target_today_str}_tiptop_xml.csv\n\n"
        f"type {target_today_str}_topcust_xml.csv\n\n"
        f"pause"
    )

    for filename, text_content in [("rpt_diff.bat", rpt_content), ("xml_diff.bat", xml_content)]:
        try:
            target_bat_path = os.path.join(aud_dir_path, filename)
            with open(target_bat_path, 'w', encoding='cp950', newline='\r\n') as f:
                f.write(text_content)
            print(f"遠端 RDP 腳本重構成功 -> {target_bat_path}", flush=True)
        except Exception as e:
            print(f"重構 Bat 腳本失敗 ({filename}): {e}", flush=True)

def merge_diff_files_to_attachment(local_dir=None, target_dir=None, target_filename=None, *args, **kwargs):
    """ 統一對齊至核心 aud 目錄：讀取與 z_chk_all.sh 相同目錄下的 4fd_diff.txt 與 4gl_diff.txt """
    t_dir = target_dir or (args[1] if len(args) > 1 else None)
    t_filename = target_filename or (args[2] if len(args) > 2 else None)

    if not t_dir or not t_filename:
        print("合併失敗：未收到正確的目標目錄或目標檔名。", flush=True)
        return

    project_root_dir = local_dir or get_project_root()
    aud_dir_path = os.path.join(project_root_dir, "aud")

    path_4fd = os.path.join(aud_dir_path, "4fd_diff.txt")
    path_4gl = os.path.join(aud_dir_path, "4gl_diff.txt")
    destination_file_path = os.path.join(t_dir, t_filename)
    content_segments = []

    print(f"第三階段：正在從本地核心 aud 資料夾讀取差異結果... 目錄: {aud_dir_path}", flush=True)

    for p in [path_4fd, path_4gl]:
        if os.path.exists(p):
            try:
                with open(p, 'r', encoding='utf-8-sig') as f:
                    c = f.read().strip()
                if c:
                    content_segments.append(c)
            except Exception as e:
                print(f"讀取本地 aud Diff 檔案失敗 {os.path.basename(p)}，嘗試改用 cp950 解碼... 錯誤: {e}", flush=True)
                try:
                    with open(p, 'r', encoding='cp950') as f:
                        c = f.read().strip()
                    if c:
                        content_segments.append(c)
                except Exception as e_inner:
                    print(f"無法讀取 Diff 檔案 {os.path.basename(p)}: {e_inner}", flush=True)
        else:
            print(f"找不到實體 Diff 檔案 (系統預期路徑): {p}", flush=True)

    if content_segments:
        try:
            with open(destination_file_path, 'w', encoding='utf-8') as wf:
                wf.write("\n\n".join(content_segments) + "\n")
            print(f"成功合併 aud 中的 4gl/4fd 內容至今日覆核資料夾 -> {t_filename}", flush=True)
        except Exception as e:
            print(f"寫入附件二失敗: {e}", flush=True)
    else:
        print(f"警告：在 {aud_dir_path} 下未找到 any 4gl_diff.txt 或 4fd_diff.txt 的內容或檔案為空。", flush=True)

def merge_csv_files_to_attachment_three(local_dir=None, target_dir=None, target_filename=None, today_str=None, *args, **kwargs):
    """ 統一對齊至核心 aud 目錄：直接在與 z_chk_all.sh 相同的 aud 目錄下尋找下載放回的 4 個 CSV 檔案 """
    t_dir = target_dir or (args[1] if len(args) > 1 else None)
    t_filename = target_filename or (args[2] if len(args) > 2 else None)
    t_today_str = today_str or (args[3] if len(args) > 3 else datetime.now().strftime("%Y%m%d"))

    if not t_dir or not t_filename:
        print("寫入附件三失敗：目標目錄或檔名無效。", flush=True)
        return

    project_root_dir = local_dir or get_project_root()
    aud_dir_path = os.path.join(project_root_dir, "aud")
    destination_file_path = os.path.join(t_dir, t_filename)

    def get_csv_block_content(suffix):
        filename = f"{t_today_str}{suffix}"
        file_path = os.path.join(aud_dir_path, filename)

        if not os.path.exists(file_path) and t_dir:
            file_path = os.path.join(t_dir, filename)

        block_str = f"D:\\aud>type {filename}\r\n"

        if os.path.exists(file_path):
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                block_str += f"{content}\r\n" if content else "\r\n"
            except Exception as e:
                print(f"讀取檔案實體失敗 {filename}，嘗試改用 CP950 編碼解碼... 錯誤原因: {e}", flush=True)
                try:
                    with open(file_path, 'r', encoding='cp950') as f:
                        content = f.read().strip()
                    block_str += f"{content}\r\n" if content else "\r\n"
                except Exception as e_inner:
                    block_str += "(讀取檔案失敗。)\r\n"
        else:
            block_str += "(系統找不到指定的檔案。)\r\n"

        return block_str

    print(f"附件三開始尋找 CSV 來源。統一綁定本地核心目錄: {aud_dir_path}", flush=True)

    final_output = get_csv_block_content("_tiptop_rpt.csv") + "\r\n"
    final_output += get_csv_block_content("_topcust_rpt.csv") + "\r\n"
    final_output += "D:\\aud>pause\r\n請按任意鍵繼續 . . .\r\n\r\n"

    final_output += get_csv_block_content("_tiptop_xml.csv") + "\r\n"
    final_output += get_csv_block_content("_topcust_xml.csv") + "\r\n"
    final_output += "D:\\aud>pause\r\n請按任意鍵繼續 . . .\r\n"

    try:
        with open(destination_file_path, 'w', encoding='utf-8', newline='\r\n') as wf:
            wf.write(final_output)
        print(f"【路徑整合成功】已自核心 aud 目錄抓到 CSV 並導出至今日覆核區 -> {t_filename}", flush=True)
    except Exception as e:
        print(f"❌ 寫入附件三完整文字檔失敗: {e}", flush=True)

def audit_and_compare_logs(excel_path=None, att_two_path=None, att_three_path=None, output_log_dir=None, today_str=None, *args, **kwargs):
    """
    【雙向精密聚焦 - 正反向燈號全方位稽核版】
    1. 正向稽核 (實體 -> Excel)：實體環境有的檔案，Excel F 欄必須有登記且對應燈號必須為 Y。
    2. 反向稽核 (Excel -> 實體)：Excel 紅色列中只要指標勾選 'Y'，實體環境就絕對必須有對應檔案。
    """
    ex_path = excel_path or (args[0] if len(args) > 0 else None)
    t_str = today_str or (kwargs.get("today_str") if "today_str" in kwargs else datetime.now().strftime("%Y%m%d"))
    a2_path = att_two_path or (args[1] if len(args) > 1 else None)
    a3_path = att_three_path or (args[2] if len(args) > 2 else None)

    project_root_dir = get_project_root()
    aud_dir_path = output_log_dir or os.path.join(project_root_dir, "aud")
    current_year_sheet_name = datetime.now().strftime("%Y")

    print(f"稽核雙向核心啟動：正在讀取本次產生的 Excel -> {ex_path}", flush=True)

    detected_4gl_files = set()
    detected_4fd_files = set()
    detected_rpt_files = set()
    detected_xml_files = set()
    all_detected_files = set()

    error_logs = []

    try:
        if a2_path and os.path.exists(str(a2_path)):
            with open(a2_path, 'r', encoding='utf-8', errors='ignore') as f:
                found = re.findall(r'[a-zA-Z0-9_\-]+\.(?:4gl|4fd)', f.read(), re.IGNORECASE)
                for item in found:
                    fname = item.lower().strip()
                    all_detected_files.add(fname)
                    if fname.endswith('.4gl'):
                        detected_4gl_files.add(fname)
                    elif fname.endswith('.4fd'):
                        detected_4fd_files.add(fname)

        if a3_path and os.path.exists(str(a3_path)):
            with open(a3_path, 'r', encoding='utf-8', errors='ignore') as f:
                found = re.findall(r'[a-zA-Z0-9_\-]+\.(?:rpt|xml)', f.read(), re.IGNORECASE)
                for item in found:
                    fname = item.lower().strip()
                    all_detected_files.add(fname)
                    if fname.endswith('.rpt'):
                        detected_rpt_files.add(fname)
                    elif fname.endswith('.xml'):
                        detected_xml_files.add(fname)
    except Exception as e:
        print(f"解析實體檔案異常: {e}", flush=True)

    excel_red_rows_data = []
    wb = None
    try:
        if ex_path and os.path.exists(str(ex_path)):
            wb = openpyxl.load_workbook(ex_path, data_only=True)
            target_sheet = None
            for sheet_name in wb.sheetnames:
                if current_year_sheet_name in sheet_name:
                    target_sheet = wb[sheet_name]
                    break

            if target_sheet:
                for row in target_sheet.iter_rows(min_row=1):
                    row_is_red = False
                    row_cells = {}

                    for cell in row:
                        if cell.coordinate:
                            col_letter = re.sub(r'\d', '', cell.coordinate).upper()
                            val_str = str(cell.value or "").strip()
                            row_cells[col_letter] = val_str

                            if is_red_color(cell):
                                row_is_red = True

                    if row_is_red and row_cells:
                        excel_red_rows_data.append(row_cells)
    except Exception as excel_err:
        print(f"讀取 Excel 發生異常: {excel_err}", flush=True)
    finally:
        if wb:
            try:
                wb.close()
            except Exception as close_err:
                print(f"關閉 Excel Workbook 發生異常: {close_err}", flush=True)

    print(f"成功讀取 Excel 紅色列資料，共取得: {len(excel_red_rows_data)} 條紀錄", flush=True)

    # 正向稽核
    for f_name in sorted(list(all_detected_files)):
        pure_base_name = os.path.splitext(f_name)[0].strip().lower()

        matched_rows = []
        for row in excel_red_rows_data:
            f_val = str(row.get('F', '')).strip()
            if f_val and is_file_matching_excel(pure_base_name, f_val):
                matched_rows.append(row)

        if not matched_rows:
            error_logs.append(f"[正向異常] 實體環境有檔: {f_name} -> 原因：Excel 當年度紅色列中完全無此登記紀錄")
            continue

        if f_name.endswith('.4gl'):
            if not any(str(r.get('I', '')).strip().upper() == 'Y' for r in matched_rows):
                last_val = str(matched_rows[-1].get('I', '')).strip().upper() or "空值"
                error_logs.append(f"[正向異常] 實體環境有檔: {f_name} -> 原因：Excel 有登記，但簽核放行指標 I 欄位目前為 '{last_val}'")

        elif f_name.endswith('.4fd'):
            if not any(str(r.get('H', '')).strip().upper() == 'Y' for r in matched_rows):
                last_val = str(matched_rows[-1].get('H', '')).strip().upper() or "空值"
                error_logs.append(f"[正向異常] 實體環境有檔: {f_name} -> 原因：Excel 有登記，但簽核放行指標 H 欄位目前為 '{last_val}'")

        elif f_name.endswith('.rpt'):
            if not any(str(r.get('J', '')).strip().upper() == 'Y' for r in matched_rows):
                last_val = str(matched_rows[-1].get('J', '')).strip().upper() or "空值"
                error_logs.append(f"[正向異常] 實體環境有檔: {f_name} -> 原因：Excel 有登記，但報表簽核指標 J 欄位目前為 '{last_val}'")

        elif f_name.endswith('.xml'):
            if not any(str(r.get('K', '')).strip().upper() == 'Y' for r in matched_rows):
                last_val = str(matched_rows[-1].get('K', '')).strip().upper() or "空值"
                error_logs.append(f"[正向異常] 實體環境有檔: {f_name} -> 原因：Excel 有登記，但規格檔簽核指標 K 欄位目前為 '{last_val}'")

    # 反向稽核
    for row in excel_red_rows_data:
        f_val = str(row.get('F', '')).strip()
        if not f_val:
            continue

        if str(row.get('I', '')).strip().upper() == 'Y':
            has_matched_4gl = any(is_file_matching_excel(os.path.splitext(f_name)[0], f_val) for f_name in detected_4gl_files)
            if not has_matched_4gl:
                expected_file = f"{f_val}.4gl"
                error_logs.append(f"[反向異常] Excel 登記放行: 程式代號 [{row.get('F')}] -> 原因：I 欄為 Y，但在附件二中找不到對應實體檔案 {expected_file}")

        if str(row.get('H', '')).strip().upper() == 'Y':
            has_matched_4fd = any(is_file_matching_excel(os.path.splitext(f_name)[0], f_val) for f_name in detected_4fd_files)
            if not has_matched_4fd:
                expected_file = f"{f_val}.4fd"
                error_logs.append(f"[反向異常] Excel 登記放行: 程式代號 [{row.get('F')}] -> 原因：H 欄為 Y，但在附件二中找不到對應實體檔案 {expected_file}")

        if str(row.get('J', '')).strip().upper() == 'Y':
            has_matched_rpt = any(is_file_matching_excel(os.path.splitext(f_name)[0], f_val) for f_name in detected_rpt_files)
            if not has_matched_rpt:
                error_logs.append(f"[反向異常] Excel 登記放行: 程式代號 [{row.get('F')}] -> 原因：J 欄為 Y，但在附件三中找不到任何以 {f_val} 開頭的實體 .rpt 報表檔案")

        if str(row.get('K', '')).strip().upper() == 'Y':
            has_matched_xml = any(is_file_matching_excel(os.path.splitext(f_name)[0], f_val) for f_name in detected_xml_files)
            if not has_matched_xml:
                expected_file = f"{f_val}.xml"
                error_logs.append(f"[反向異常] Excel 登記放行: 程式代號 [{row.get('F')}] -> 原因：K 欄為 Y，但在附件三中找不到對應實體檔案 {expected_file}")

    os.makedirs(aud_dir_path, exist_ok=True)
    report_file_path = os.path.join(aud_dir_path, f"{t_str}_稽核比對結果報告.txt")

    try:
        with open(report_file_path, 'w', encoding='utf-8') as rf:
            rf.write("====================================================\n")
            rf.write(f"程式修改定期稽核雙向異常回報 ({t_str})\n")
            rf.write("====================================================\n\n")
            rf.write(f"工作表頁籤: {current_year_sheet_name} 年度\n")
            rf.write(f"比對來源檔案: {os.path.basename(str(ex_path)) if ex_path else '無'}\n\n")

            if error_logs:
                rf.write(f"稽核結論: 雙向檢驗不通過！共偵測到 {len(error_logs)} 筆正反向不對齊異常！\n\n")
                rf.write("---------------- [ 以下為異常錯誤清單 ] ----------------\n")
                for item in error_logs:
                    rf.write(f"{item}\n")
            else:
                rf.write("稽核結論: 恭喜！本週 [正向:實體->Excel] 與 [反向:Excel->實體] 雙向 100% 完全對齊，無任何異常！\n")

            rf.write("\n====================================================\n")

        print(f"雙向稽核報告已成功寫入磁碟 -> {report_file_path}", flush=True)
    except Exception as e:
        print(f"實體寫入報告失敗: {e}", flush=True)

    return True

def deploy_local_folder_to_network_share(local_folder_path, remote_base_dir):
    print(f"開始將本機成果同步至網碟發布目錄...", flush=True)
    if not local_folder_path or not os.path.exists(local_folder_path):
        print("同步失敗：本地工作區路徑不存在。", flush=True)
        return

    folder_name = os.path.basename(local_folder_path)
    remote_target_path = os.path.join(remote_base_dir, folder_name)
    try:
        if os.path.exists(remote_target_path):
            print("遠端網碟已存在同名資料夾，執行覆蓋同步......", flush=True)
        os.makedirs(remote_target_path, exist_ok=True)

        for root, dirs, files in os.walk(local_folder_path):
            rel_path = os.path.relpath(root, local_folder_path)
            dest_dir = remote_target_path if rel_path == "." else os.path.join(remote_target_path, rel_path)
            os.makedirs(dest_dir, exist_ok=True)
            for file in files:
                shutil.copy2(os.path.join(root, file), os.path.join(dest_dir, file))
        print(f"網碟同步成功！目標位置: {remote_target_path}", flush=True)
    except Exception as e:
        print(f"網碟同步失敗（請檢查網路連線或網碟存取權限）: {e}", flush=True)
